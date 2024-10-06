import pandas as pd
import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import os
from PIL import Image as PilImage
from openpyxl.styles import Alignment, Font
from datetime import datetime

# Step 1: Read settings from a settings.txt file
def read_settings(settings_file):
    settings = {}
    try:
        with open(settings_file, 'r') as file:
            for line in file:
                if line.strip() and not line.startswith("#"):
                    key, value = line.strip().split('=', 1)
                    settings[key.strip()] = value.strip()
    except FileNotFoundError:
        print(f"Error: The settings file '{settings_file}' does not exist.")
        return None
    except Exception as e:
        print(f"Error reading settings: {e}")
        return None
    
    return settings

# Step 2: Read Excel file
def read_excel(file_path, url_column_name):
    # Read Excel file using pandas
    df = pd.read_excel(file_path)
    
    # Ensure the URL column exists
    if url_column_name not in df.columns:
        raise ValueError(f"Column '{url_column_name}' not found in the Excel file.")
    
    return df

# Step 3: Generate QR code (without AgentCode)
def generate_qr_code(url, size=(150, 150)):
    # Generate QR code with higher error correction and fit for long URLs
    qr = qrcode.QRCode(
        version=5,  # Adjust for long URLs, version 5 can handle up to 114 characters
        error_correction=qrcode.constants.ERROR_CORRECT_H,  # High error correction
        box_size=10,  # Larger box size for better readability
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    
    # Create an image from the QR code
    img = qr.make_image(fill='black', back_color='white')
    
    # Convert to PIL Image for editing
    pil_img = img.convert("RGB")
    pil_img = pil_img.resize(size)

    # Save the image to a BytesIO stream
    byte_io = BytesIO()
    pil_img.save(byte_io, format='PNG')
    
    return byte_io

# Step 4: Update Excel with QR codes in the new 'QrCode' column
def update_excel_with_qr_codes(excel_path, url_column_name, output_excel_path):
    # Read the Excel file using pandas and openpyxl
    df = read_excel(excel_path, url_column_name)
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Add the 'QrCode' column if it doesn't exist
    max_column = ws.max_column + 1
    qrcode_column = chr(ord('A') + max_column - 1)
    ws.cell(row=1, column=max_column, value="QrCode")
    
    # Match the 'QrCode' header design with 'URL' header design
    url_header_cell = ws[f"A1"]  # Assuming 'URL' header is in the first column (A)
    qr_code_header_cell = ws[f'{qrcode_column}1']
    
    # Manually copy font and alignment from 'URL' header to 'QrCode' header
    qr_code_header_cell.font = Font(name=url_header_cell.font.name, 
                                    bold=url_header_cell.font.bold,
                                    size=url_header_cell.font.size, 
                                    color=url_header_cell.font.color)
    
    qr_code_header_cell.alignment = Alignment(horizontal=url_header_cell.alignment.horizontal,
                                              vertical=url_header_cell.alignment.vertical)

    # Set the size of the new column and rows to fit the QR code images
    qr_code_img_size = 150  # Set the size of the QR code (150x150 pixels)
    column_width = qr_code_img_size / 7  # Approximation of pixels to Excel column width
    row_height = qr_code_img_size / 0.75  # Approximation of pixels to Excel row height
    
    ws.column_dimensions[qrcode_column].width = column_width  # Adjust the column width for the QR code
    for row in range(2, len(df) + 2):
        ws.row_dimensions[row].height = row_height  # Adjust row height to fit the QR code

    # Wrap text in the 'URL' column (Assuming it's column A)
    for cell in ws['A']:
        cell.alignment = Alignment(wrap_text=True)

    # Track the total number of records processed and QR codes generated
    total_records = len(df)
    qr_codes_created = 0

    # Iterate over the rows of the dataframe
    for index, row in df.iterrows():
        url = row[url_column_name]
        
        # Generate QR code for each URL
        qr_code_img = generate_qr_code(url, size=(qr_code_img_size, qr_code_img_size))  # Adjust the size for better readability
        
        # Convert the BytesIO image into an Image object for openpyxl
        img = Image(qr_code_img)
        
        # Align the QR code image in the center of the cell
        img.anchor = f'{qrcode_column}{index + 2}'  # Insert QR code in the new 'QrCode' column
        ws[f'{qrcode_column}{index + 2}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add image to the Excel sheet
        ws.add_image(img)
        qr_codes_created += 1
    
    # Save the updated Excel file
    wb.save(output_excel_path)

    # Return the total number of records and QR codes created
    return total_records, qr_codes_created

# Step 5: Main function with settings from a file
def main():
    # Path to the settings file
    settings_file = "settings.txt"

    # Read settings
    settings = read_settings(settings_file)
    if settings is None:
        return

    # Get important settings from the file
    input_file = settings.get("input_file")
    url_column_name = settings.get("url_column_name")
    output_file = settings.get("output_file")

    # Check if all required settings are present
    if not input_file or not url_column_name or not output_file:
        print("Error: Missing required settings. Please check your settings.txt file.")
        return
    
    # Ensure the input file exists
    if not os.path.isfile(input_file):
        print(f"Error: The file '{input_file}' does not exist.")
        return

    # Append the current date and time to the output file name
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file_with_timestamp = f"{output_file.rstrip('.xlsx')}_{current_datetime}.xlsx"

    try:
        # Process the Excel file and generate QR codes
        total_records, qr_codes_created = update_excel_with_qr_codes(input_file, url_column_name, output_file_with_timestamp)
        
        # Display summary of results
        print(f"\nProcess Completed:")
        print(f"Total records read: {total_records}")
        print(f"Total QR codes created: {qr_codes_created}")
        print(f"QR codes have been generated and saved to '{output_file_with_timestamp}' successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

    # Wait for the user to press any key before closing the console
    input("\nPress any key to close the console...")

# Entry point for the script
if __name__ == "__main__":
    main()
